"""
Geração do relatório de conciliação em Excel.

O entregável para o contador é a planilha, não o terminal. A aba "Resumo"
existe para ser lida em 30 segundos; as demais servem à auditoria do que
foi decidido pelo motor.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AZUL = "1F3864"
CINZA = "F2F2F2"
VERMELHO = "FFC7CE"
AMARELO = "FFEB9C"
VERDE = "C6EFCE"

ROTULOS = {
    "nao_contabilizado": "Nao contabilizado (banco sem razao)",
    "nao_compensado": "Nao compensado (razao sem banco)",
    "divergencia_valor": "Divergencia de valor",
    "divergencia_data": "Divergencia de data (competencia)",
}


def _estilizar_cabecalho(ws, n_colunas: int) -> None:
    fill = PatternFill("solid", fgColor=AZUL)
    for col in range(1, n_colunas + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autoajustar(ws, df: pd.DataFrame, largura_max: int = 46) -> None:
    for i, coluna in enumerate(df.columns, start=1):
        conteudo = df[coluna].astype(str).str.len().max() if len(df) else 0
        largura = min(max(int(conteudo or 0), len(str(coluna))) + 3, largura_max)
        ws.column_dimensions[get_column_letter(i)].width = largura


def gerar_excel(resultado: dict, caminho: str | Path) -> Path:
    """Escreve o relatório completo e devolve o caminho do arquivo."""
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    resumo = resultado["resumo"]
    divergencias = resultado["divergencias"].copy()
    conciliados = resultado["conciliados"].copy()

    if not divergencias.empty:
        divergencias.insert(1, "descricao", divergencias["tipo"].map(ROTULOS))

    painel = pd.DataFrame(
        [
            ("Linhas no extrato", resumo["linhas_extrato"]),
            ("Linhas no razao", resumo["linhas_razao"]),
            ("Conciliados automaticamente", resumo["conciliados"]),
            ("Taxa de conciliacao", f"{resumo['taxa_conciliacao']:.2%}"),
            ("Pendentes no extrato", resumo["pendentes_extrato"]),
            ("Pendentes no razao", resumo["pendentes_razao"]),
            ("Divergencias criticas", resumo["divergencias_criticas"]),
            ("Valor em aberto (R$)", f"{resumo['valor_em_aberto']:,.2f}"),
        ],
        columns=["Indicador", "Valor"],
    )

    if not divergencias.empty:
        por_tipo = (
            divergencias.groupby("tipo")
            .agg(ocorrencias=("valor", "size"), valor_total=("valor", lambda s: s.abs().sum()))
            .reset_index()
        )
        por_tipo["descricao"] = por_tipo["tipo"].map(ROTULOS)
        por_tipo = por_tipo[["descricao", "ocorrencias", "valor_total"]]
    else:
        por_tipo = pd.DataFrame(columns=["descricao", "ocorrencias", "valor_total"])

    metodos = pd.DataFrame(
        sorted(resumo["por_metodo"].items(), key=lambda kv: -kv[1]),
        columns=["metodo", "pares"],
    )

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        painel.to_excel(writer, sheet_name="Resumo", index=False, startrow=0)
        por_tipo.to_excel(writer, sheet_name="Resumo", index=False, startrow=len(painel) + 3)
        divergencias.to_excel(writer, sheet_name="Divergencias", index=False)
        conciliados.to_excel(writer, sheet_name="Conciliados", index=False)
        metodos.to_excel(writer, sheet_name="Metodos", index=False)

        wb = writer.book

        ws = wb["Resumo"]
        _estilizar_cabecalho(ws, 2)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 20
        linha_titulo = len(painel) + 4
        for col in range(1, 4):
            c = ws.cell(row=linha_titulo, column=col)
            c.font = Font(color="FFFFFF", bold=True)
            c.fill = PatternFill("solid", fgColor=AZUL)
        ws.column_dimensions["C"].width = 18

        ws = wb["Divergencias"]
        _estilizar_cabecalho(ws, max(len(divergencias.columns), 1))
        _autoajustar(ws, divergencias)
        if not divergencias.empty:
            col_sev = list(divergencias.columns).index("severidade") + 1
            for linha in range(2, len(divergencias) + 2):
                sev = ws.cell(row=linha, column=col_sev).value
                cor = VERMELHO if sev == "alta" else AMARELO
                for col in range(1, len(divergencias.columns) + 1):
                    ws.cell(row=linha, column=col).fill = PatternFill("solid", fgColor=cor)

        for nome, df in (("Conciliados", conciliados), ("Metodos", metodos)):
            ws = wb[nome]
            _estilizar_cabecalho(ws, max(len(df.columns), 1))
            _autoajustar(ws, df)

    return caminho


def gerar_csv(resultado: dict, pasta: str | Path) -> dict:
    """Alternativa leve: dois CSVs para quem prefere importar no ERP."""
    pasta = Path(pasta)
    pasta.mkdir(parents=True, exist_ok=True)
    caminhos = {
        "divergencias": pasta / "divergencias.csv",
        "conciliados": pasta / "conciliados.csv",
    }
    resultado["divergencias"].to_csv(caminhos["divergencias"], index=False)
    resultado["conciliados"].to_csv(caminhos["conciliados"], index=False)
    return caminhos
